import os
import xlwt
import openpyxl

# Ruta de la carpeta que contiene los archivos
carpeta = '/home/roadcode/Music'

# Lista para almacenar los nombres de archivo
archivos = []

# Recorre los archivos en la carpeta
for archivo in os.listdir(carpeta):
    if os.path.isfile(os.path.join(carpeta, archivo)):
        archivos.append(archivo)

# Exportar a un archivo de texto (TXT)
with open('archivos.txt', 'w') as archivo_txt:
    for archivo in archivos:
        archivo_txt.write(archivo + '\n')

# Exportar a un archivo de Excel (XLS)
archivo_xls = xlwt.Workbook()
hoja_xls = archivo_xls.add_sheet('Archivos')
for i, archivo in enumerate(archivos):
    hoja_xls.write(i, 0, archivo)
archivo_xls.save('archivos.xls')

# Exportar a un archivo de Excel (XLSX)
archivo_xlsx = openpyxl.Workbook()
hoja_xlsx = archivo_xlsx.active
for i, archivo in enumerate(archivos):
    hoja_xlsx.cell(row=i+1, column=1, value=archivo)
archivo_xlsx.save('archivos.xlsx')

print('Archivos exportados a archivos.txt, archivos.xls y archivos.xlsx')
